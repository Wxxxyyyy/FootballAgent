# -*- coding: utf-8 -*-
"""
2026 世界杯国家队历史交锋数据导入 Neo4j

数据来源: football-data.co.uk 的 WorldCup2026.xlsx
  - WorldCup2026: 2026世界杯正赛（进行中）
  - WorldCup2022: 2022卡塔尔世界杯
  - WorldCup2018: 2018俄罗斯世界杯
  - WorldCup2014: 2014巴西世界杯
  - WorldCup2026Qualifiers: 2026预选赛

导入内容:
  - NationalTeam 节点（和联赛 Club Team 分开，用不同标签）
  - PLAYED_AGAINST 关系（HomeTeam → AwayTeam）
  - 关系属性: match_date, season, match_result, total_goals, competition, odds_info

用法:
  # 先启动 Neo4j 容器
  docker-compose up -d neo4j

  # 导入国家队数据
  python -m pipeline.national_team_neo4j_loader
"""

import os
import sys
import logging
import httpx
import openpyxl
import pandas as pd
from pathlib import Path
from datetime import datetime
from typing import Optional

from dotenv import load_dotenv

PROJECT_ROOT = Path(__file__).resolve().parent.parent
load_dotenv(os.path.join(PROJECT_ROOT, ".env"))

logger = logging.getLogger(__name__)

# Neo4j 配置
NEO4J_URI = os.getenv("NEO4J_URL", "bolt://localhost:7687")
NEO4J_USER = os.getenv("NEO4J_USERNAME", "neo4j")
NEO4J_PASSWORD = os.getenv("NEO4J_PASSWORD", "football123")
NEO4J_DATABASE = os.getenv("NEO4J_DATABASE", "neo4j")

# WorldCup2026.xlsx 下载地址
WC_XLSX_URL = "https://www.football-data.co.uk/WorldCup2026.xlsx"
WC_XLSX_PATH = os.path.join(PROJECT_ROOT, "data", "WorldCup2026.xlsx")

# ═══════════════════════════════════════════════════════════════
#  2026 世界杯 48 支参赛队
# ═══════════════════════════════════════════════════════════════

# 中文 → 英文映射（xlsx 中用英文名）
# 同时包含可能的别名，用于匹配 xlsx 中的队名
NATIONAL_TEAMS = {
    # A组
    "墨西哥": ["Mexico"],
    "南非": ["South Africa"],
    "韩国": ["South Korea", "Korea Republic", "Korea"],
    "捷克": ["Czech Republic", "Czechia", "Czech"],
    # B组
    "加拿大": ["Canada"],
    "波黑": ["Bosnia & Herzegovina", "Bosnia and Herzegovina", "Bosnia"],
    "卡塔尔": ["Qatar"],
    "瑞士": ["Switzerland"],
    # C组
    "巴西": ["Brazil"],
    "摩洛哥": ["Morocco"],
    "海地": ["Haiti"],
    "苏格兰": ["Scotland"],
    # D组
    "美国": ["USA", "United States"],
    "巴拉圭": ["Paraguay"],
    "澳大利亚": ["Australia"],
    "土耳其": ["Turkey"],
    # E组
    "德国": ["Germany"],
    "库拉索": ["Curacao", "Curaçao"],
    "科特迪瓦": ["Ivory Coast", "Côte d'Ivoire", "Cote d'Ivoire"],
    "厄瓜多尔": ["Ecuador"],
    # F组
    "荷兰": ["Netherlands", "Holland"],
    "日本": ["Japan"],
    "瑞典": ["Sweden"],
    "突尼斯": ["Tunisia"],
    # G组
    "比利时": ["Belgium"],
    "埃及": ["Egypt"],
    "伊朗": ["Iran"],
    "新西兰": ["New Zealand"],
    # H组
    "西班牙": ["Spain"],
    "佛得角": ["Cape Verde", "Cabo Verde"],
    "沙特阿拉伯": ["Saudi Arabia"],
    "乌拉圭": ["Uruguay"],
    # I组
    "法国": ["France"],
    "塞内加尔": ["Senegal"],
    "伊拉克": ["Iraq"],
    "挪威": ["Norway"],
    # J组
    "阿根廷": ["Argentina"],
    "阿尔及利亚": ["Algeria"],
    "奥地利": ["Austria"],
    "约旦": ["Jordan"],
    # K组
    "葡萄牙": ["Portugal"],
    "刚果（金）": ["DR Congo", "Congo DR", "Democratic Republic of Congo", "Congo (Kinshasa)"],
    "乌兹别克斯坦": ["Uzbekistan"],
    "哥伦比亚": ["Colombia"],
    # L组
    "英格兰": ["England"],
    "克罗地亚": ["Croatia"],
    "加纳": ["Ghana"],
    "巴拿马": ["Panama"],
}

# 组别映射
GROUP_MAP = {
    "Mexico": "A", "South Africa": "A", "South Korea": "A", "Czech Republic": "A",
    "Canada": "B", "Bosnia & Herzegovina": "B", "Qatar": "B", "Switzerland": "B",
    "Brazil": "C", "Morocco": "C", "Haiti": "C", "Scotland": "C",
    "USA": "D", "Paraguay": "D", "Australia": "D", "Turkey": "D",
    "Germany": "E", "Curacao": "E", "Ivory Coast": "E", "Ecuador": "E",
    "Netherlands": "F", "Japan": "F", "Sweden": "F", "Tunisia": "F",
    "Belgium": "G", "Egypt": "G", "Iran": "G", "New Zealand": "G",
    "Spain": "H", "Cape Verde": "H", "Saudi Arabia": "H", "Uruguay": "H",
    "France": "I", "Senegal": "I", "Iraq": "I", "Norway": "I",
    "Argentina": "J", "Algeria": "J", "Austria": "J", "Jordan": "J",
    "Portugal": "K", "DR Congo": "K", "Uzbekistan": "K", "Colombia": "K",
    "England": "L", "Croatia": "L", "Ghana": "L", "Panama": "L",
}

# 构建英文→中文反向映射（取每个中文名的第一个英文别名作为标准名）
EN_TO_ZH = {}
STANDARD_EN_NAMES = set()
for zh, en_list in NATIONAL_TEAMS.items():
    standard_en = en_list[0]  # 第一个作为标准英文名
    STANDARD_EN_NAMES.add(standard_en)
    EN_TO_ZH[standard_en] = zh
    for alias in en_list:
        EN_TO_ZH[alias.lower()] = zh


def normalize_team_name(name: str) -> Optional[str]:
    """将 xlsx 中的队名标准化为标准英文名"""
    if not name:
        return None
    name = str(name).strip()
    # 精确匹配
    for zh, en_list in NATIONAL_TEAMS.items():
        if name in en_list:
            return en_list[0]  # 返回标准英文名
    # 模糊匹配（忽略大小写）
    name_lower = name.lower()
    for zh, en_list in NATIONAL_TEAMS.items():
        for alias in en_list:
            if alias.lower() == name_lower:
                return en_list[0]
    return None


# ═══════════════════════════════════════════════════════════════
#  下载 WorldCup2026.xlsx
# ═══════════════════════════════════════════════════════════════

def download_wc_xlsx() -> str:
    """下载 WorldCup2026.xlsx 到本地"""
    if os.path.exists(WC_XLSX_PATH):
        logger.info(f"已存在: {WC_XLSX_PATH}")
        return WC_XLSX_PATH

    logger.info(f"下载 {WC_XLSX_URL}...")
    headers = {"User-Agent": "Mozilla/5.0"}
    resp = httpx.get(WC_XLSX_URL, headers=headers, timeout=30, follow_redirects=True)
    if resp.status_code != 200:
        raise RuntimeError(f"下载失败: {resp.status_code}")

    os.makedirs(os.path.dirname(WC_XLSX_PATH), exist_ok=True)
    with open(WC_XLSX_PATH, "wb") as f:
        f.write(resp.content)
    logger.info(f"下载完成: {len(resp.content)} bytes → {WC_XLSX_PATH}")
    return WC_XLSX_PATH


# ═══════════════════════════════════════════════════════════════
#  读取比赛数据
# ═══════════════════════════════════════════════════════════════

def extract_matches(xlsx_path: str) -> list[dict]:
    """
    从 WorldCup2026.xlsx 提取涉及48支参赛队的比赛

    返回: [{"home", "away", "date", "home_goals", "away_goals", "result",
            "competition", "season", "odds"}, ...]
    """
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    matches = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        headers = [str(h) if h else "" for h in rows[0]]

        # 确定列索引（不同工作表列名不同）
        col_map = {}
        for i, h in enumerate(headers):
            hl = h.lower()
            if h == "Home" or hl == "hometeam":
                col_map["home"] = i
            elif h == "Away" or hl == "awayteam":
                col_map["away"] = i
            elif h == "Date":
                col_map["date"] = i
            elif h == "HGFT" or h == "HG" or h == "FTHG":
                col_map["hg"] = i
            elif h == "AGFT" or h == "AG" or h == "FTAG":
                col_map["ag"] = i
            elif h == "Competition":
                col_map["competition"] = i
            elif h == "bet365-H" or h == "B365H":
                col_map["b365h"] = i
            elif h == "bet365-D" or h == "B365D":
                col_map["b365d"] = i
            elif h == "bet365-A" or h == "B365A":
                col_map["b365a"] = i

        if "home" not in col_map or "away" not in col_map:
            continue

        # 确定赛季
        if "2026" in sheet_name and "Qual" not in sheet_name:
            season = "2026"
        elif "2026" in sheet_name:
            season = "2026-Q"
        elif "2022" in sheet_name:
            season = "2022"
        elif "2018" in sheet_name:
            season = "2018"
        elif "2014" in sheet_name:
            season = "2014"
        else:
            season = sheet_name

        for row in rows[1:]:
            try:
                home_raw = row[col_map["home"]] if "home" in col_map else None
                away_raw = row[col_map["away"]] if "away" in col_map else None
                if not home_raw or not away_raw:
                    continue

                home = normalize_team_name(str(home_raw))
                away = normalize_team_name(str(away_raw))

                # 至少有一方是48支参赛队
                if not home and not away:
                    continue

                # 如果有一方不是参赛队，也跳过（只导入参赛队之间的比赛）
                if not home or not away:
                    continue

                hg = row[col_map["hg"]] if "hg" in col_map else None
                ag = row[col_map["ag"]] if "ag" in col_map else None

                # 解析比分
                try:
                    hg = int(hg) if hg is not None else None
                    ag = int(ag) if ag is not None else None
                except (ValueError, TypeError):
                    hg, ag = None, None

                # 计算结果
                result = None
                if hg is not None and ag is not None:
                    if hg > ag:
                        result = "H"
                    elif hg < ag:
                        result = "A"
                    else:
                        result = "D"

                # 日期
                date_raw = row[col_map["date"]] if "date" in col_map else None
                date_str = str(date_raw)[:10] if date_raw else ""

                # 赛事
                competition = ""
                if "competition" in col_map:
                    comp_raw = row[col_map["competition"]]
                    competition = str(comp_raw) if comp_raw else sheet_name
                else:
                    competition = sheet_name

                # 赔率
                odds = {}
                for key in ["b365h", "b365d", "b365a"]:
                    if key in col_map:
                        val = row[col_map[key]]
                        if val:
                            try:
                                odds[key] = float(val)
                            except (ValueError, TypeError):
                                pass

                if hg is not None and ag is not None:
                    matches.append({
                        "home": home,
                        "away": away,
                        "date": date_str,
                        "home_goals": hg,
                        "away_goals": ag,
                        "result": result,
                        "total_goals": hg + ag,
                        "competition": competition,
                        "season": season,
                        "odds": odds if odds else None,
                    })
            except Exception as e:
                continue

    wb.close()
    logger.info(f"提取到 {len(matches)} 场涉及48支参赛队的比赛")
    return matches


# ═══════════════════════════════════════════════════════════════
#  导入 Neo4j
# ═══════════════════════════════════════════════════════════════

def import_to_neo4j(matches: list[dict]):
    """将国家队比赛数据导入 Neo4j"""

    from neo4j import GraphDatabase

    driver = GraphDatabase.driver(NEO4J_URI, auth=(NEO4J_USER, NEO4J_PASSWORD))

    try:
        with driver.session(database=NEO4J_DATABASE) as session:
            # 1. 创建唯一约束（防止重复）
            session.run(
                "CREATE CONSTRAINT IF NOT EXISTS FOR (t:NationalTeam) "
                "REQUIRE t.name IS UNIQUE"
            )
            logger.info("创建 NationalTeam 唯一约束")

            # 2. 创建所有参赛队节点
            teams_created = 0
            for zh_name, en_list in NATIONAL_TEAMS.items():
                standard_en = en_list[0]
                group = GROUP_MAP.get(standard_en, "")
                session.run(
                    "MERGE (t:NationalTeam {name: $name}) "
                    "SET t.name_zh = $zh, t.group = $group, t.type = 'national'",
                    name=standard_en, zh=zh_name, group=group,
                )
                teams_created += 1
            logger.info(f"创建 {teams_created} 个国家队节点")

            # 3. 创建比赛关系
            rel_count = 0
            for m in matches:
                # 判断结果描述
                result_text = f"{m['home_goals']}-{m['away_goals']}"
                if m["result"] == "H":
                    result_desc = f"{m['home']}胜"
                elif m["result"] == "A":
                    result_desc = f"{m['away']}胜"
                else:
                    result_desc = "平局"

                odds_json = str(m["odds"]) if m["odds"] else ""

                session.run(
                    """
                    MATCH (h:NationalTeam {name: $home}),
                          (a:NationalTeam {name: $away})
                    MERGE (h)-[r:PLAYED_AGAINST {
                        match_date: $date,
                        season: $season
                    }]->(a)
                    SET r.competition = $competition,
                        r.match_result = $result,
                        r.result_desc = $result_desc,
                        r.home_goals = $hg,
                        r.away_goals = $ag,
                        r.total_goals = $total,
                        r.odds_info = $odds
                    """,
                    home=m["home"],
                    away=m["away"],
                    date=m["date"],
                    season=m["season"],
                    competition=m["competition"],
                    result=result_text,
                    result_desc=result_desc,
                    hg=m["home_goals"],
                    ag=m["away_goals"],
                    total=m["total_goals"],
                    odds=odds_json,
                )
                rel_count += 1

            logger.info(f"创建 {rel_count} 条交锋关系")

            # 4. 统计验证
            team_count = session.run(
                "MATCH (t:NationalTeam) RETURN COUNT(t) as c"
            ).single()["c"]
            rel_count_db = session.run(
                "MATCH ()-[r:PLAYED_AGAINST]->() RETURN COUNT(r) as c"
            ).single()["c"]
            logger.info(f"Neo4j 验证: {team_count} 个国家队, {rel_count_db} 条交锋关系")

            # 显示部分数据
            sample = session.run(
                "MATCH (h:NationalTeam)-[r:PLAYED_AGAINST]->(a:NationalTeam) "
                "RETURN h.name as home, a.name as away, r.match_result as result, "
                "r.season as season, r.competition as comp "
                "ORDER BY r.match_date DESC LIMIT 10"
            )
            print("\n=== 最近10场比赛 ===")
            for r in sample:
                print(f"  {r['season']} | {r['home']} {r['result']} {r['away']} | {r['comp']}")

    finally:
        driver.close()


# ═══════════════════════════════════════════════════════════════
#  主函数
# ═══════════════════════════════════════════════════════════════

def main():
    logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")

    print("=" * 60)
    print("  2026 世界杯国家队交锋数据导入 Neo4j")
    print("=" * 60)

    # 1. 下载 xlsx
    xlsx_path = download_wc_xlsx()

    # 2. 提取比赛数据
    print("\n[1/3] 提取比赛数据...")
    matches = extract_matches(xlsx_path)
    print(f"  提取到 {len(matches)} 场比赛")

    # 按赛季统计
    from collections import Counter
    season_counts = Counter(m["season"] for m in matches)
    for season, count in sorted(season_counts.items()):
        print(f"    {season}: {count} 场")

    if not matches:
        print("❌ 未提取到比赛数据")
        return

    # 3. 导入 Neo4j
    print(f"\n[2/3] 连接 Neo4j ({NEO4J_URI})...")
    try:
        import_to_neo4j(matches)
        print(f"\n[3/3] 导入完成!")
    except Exception as e:
        print(f"❌ Neo4j 导入失败: {e}")
        print(f"  请确认 Neo4j 已启动: docker-compose up -d neo4j")
        print(f"  .env 配置: NEO4J_URL={NEO4J_URI}, NEO4J_USER={NEO4J_USER}")

    print("=" * 60)


if __name__ == "__main__":
    main()
